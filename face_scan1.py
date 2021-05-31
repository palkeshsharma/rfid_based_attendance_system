import face_camera
from cv2 import *
import mysql.connector
import face_recognition
import os
import glob
import stat
import time
import sys
from datetime import date,datetime,timedelta

import numpy as np

shiva_image = face_recognition.load_image_file(r'E:\MAJOR8\s5.jpg')
shiva_face_encoding = face_recognition.face_encodings(shiva_image)[0]


known_face_encodings = [
    shiva_face_encoding
    ]

known_face_roll = [
    "CAAA7E3C"
    ]

#Intitializing variables
face_locations = []
face_encodings = []
face_roll = []
process_this_frame = True


# initialize the camera
cam = VideoCapture(0)   # 0 -> index of camera
ret, frame = cam.read()
# resize vedio to 1/4 size
small_frame = cv2.resize(frame, (0,0), fx=0.25, fy=0.25)
#bgr to rgb
rgb_small_frame = small_frame[:,:,::-1]

#only process every other frame to save time
if process_this_frame:
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
    roll = []
    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        num = "not match"
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            num= known_face_roll[best_match_index]
            roll.append(num)
    process_this_frame = not process_this_frame
            

#display result
for (t,r,b,l), num in zip(face_locations, roll):
    t*=4
    r*=4
    b*=4
    l*=4

    cv2.rectangle(frame, (l,t),(r,b),(0,0,255),2)
    font= cv2.FONT_HERSHEY_DUPLEX
    cv2.putText(frame, num, (l+6,b-6),font , 1.0,(255,255,255),1)
    cv2.imshow("frame",frame)
    cv2.waitKey(2000)
    fetchedlnum=num

cam.release()
cv2.destroyAllWindows()

if fetchedlnum!=None:
    print(num)

else:
    print("no id for face found!!!")
    
#YAHA KA DEKHNA BACHA HE

    
import scan_id_frontend
from ardino import dat
print(dat)

rollno=dat

import datetime
import openpyxl as o
from openpyxl.styles import Font, colors



path=r"E:\MAJOR8\attendance.xlsx"
wb = o.load_workbook(path)
sheet = wb.active

if(rollno.rstrip()==fetchedlnum):
    d = datetime.datetime.now()
    last_col=sheet.max_column

    last=1
    for row in sheet:
        for cell in row:
            if cell.value==d.strftime("%x"):
                last=0
                cordinat=cell.column
                #print(cordinat)
                break
            elif cell.value is None:
                cell.value=(d.strftime("%x"))
                cordinat=cell.column
                last=0
                break
        break
            
    if last!=0:
        c1 = sheet.cell(row= 1, column=last_col+1)
        c1.value=(d.strftime("%x"))
        cordinat=cell.column
 
    for row in sheet['A']:
        if row.value == rollno.rstrip():
            #'CAAA7E3C'
            pre = sheet.cell(row=row.row, column=cordinat)
            pre.value='P'
            print('Attendance marked at '+ str(row.row) + str(cordinat))    
else:
    print("face and id not matched!!!")    
    
wb.save(r"E:\MAJOR8\attendance.xlsx")       
